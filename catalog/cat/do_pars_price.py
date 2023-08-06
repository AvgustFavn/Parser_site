
import base64
import csv
import imghdr
import io
import os
import sys
# django_project_root = 'C:\\Users\\avgus\\Documents\\py_shit\\catalog_site\\catalog'
#
# # Добавляем этот путь в sys.path
# sys.path.append(django_project_root)


import django
from django.conf import settings
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(BASE_DIR)
sys.path.append(BASE_DIR)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'cat.settings')
django.setup()

from add_prod_module import *
import subprocess
import time
import asyncio
from datetime import datetime
import xlwings as xw
from django.http import HttpRequest
from openpyxl.drawing.image import Image as XlsxImage
import openpyxl
import pandas as pd
import xlrd
from openpyxl.drawing.image import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from back import chat_get_chars

parameters = sys.argv[1:]

def pars_file(name_f, prov, user):
    print('ЗАШЛИ В ФУНКЦИЮ ПАРСА ПОСЕ ФАЙЛА')
    provider = Providers.objects.get(id=prov)
    file_path = os.path.join(BASE_DIR, 'static/files', name_f).replace('\\', '/')
    count = 0
    d = {'артикул': provider.col_art, "имя товара": provider.col_name_prod, "опт макс": provider.col_retail_price,
         "опт мин": provider.col_wholesale_price,
         "остаток на складе": provider.col_edge, "остаток Москва": provider.col_edge_mosc,
         "остаток Иркутск": provider.col_edge_irk, "остаток Новос": provider.col_edge_novos,
         'описание': provider.col_descr, 'картинка': provider.col_pic, 'скидка': provider.col_sale,
         'свободно': provider.col_free
         }


    if 'xlsx' in name_f:
        workbook = load_workbook(filename=file_path)
        sheets = workbook.sheetnames
        for sheet_name in sheets:
            sheet = workbook[sheet_name]
            for i in range(1, sheet.max_row + 1):

                # for i in range(105, 107):
                print(f'Парсим новый товар номером {i}')
                try:
                    name = sheet[f'{d["имя товара"]}{i}'].value
                except:
                    continue

                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                if name == None or name == 'имя товара':
                    continue

                try:
                    art = sheet[f'{d["артикул"]}{i}'].value
                except:
                    art = 'Н.Д.'

                try:
                    descr = sheet[f'{d["описание"]}{i}'].value
                except:
                    descr = 'Н.Д.'

                try:
                    min_ = price_validator(sheet[f'{d["опт мин"]}{i}'].value)
                    try:
                        min_ = min_ - ((min_ / 100) * price_validator(sheet[f'{d["скидка"]}{i}'].value))
                        if not min_:
                            min_ = price_validator(sheet[f'{d["опт мин"]}{i}'].value)
                    except:
                        pass
                except:
                    min_ = 'Н.Д.'

                try:
                    max_ = price_validator(sheet[f'{d["опт макс"]}{i}'].value)
                    try:
                        max_ = max_ - ((max_ / 100) * price_validator(sheet[f'{d["скидка"]}{i}'].value))
                        if not max_:
                            max_ = price_validator(sheet[f'{d["опт макс"]}{i}'].value)
                    except:
                        pass

                except:
                    max_ = 'Н.Д.'

                status = None
                edge_m = None
                edge_n = None
                edge_i = None
                try:
                    edge = sheet[f'{d["остаток на складе"]}{i}'].value
                    if edge and int(price_validator(edge)) > 0:
                        status = 'В наличии'
                    else:
                        edge = sheet[f'{d["свободно"]}{i}'].value
                        print(f'Свободные {edge}')
                        if edge and int(price_validator(edge)) > 0:
                            status = 'В наличии'
                        else:
                            status = ''
                            comm = None
                            try:
                                edge_m = sheet[f'{d["остаток Москва"]}{i}'].value
                            except:
                                edge_m = None
                            try:
                                edge_n = sheet[f'{d["остаток Новос"]}{i}'].value
                            except:
                                edge_n = None
                            try:
                                edge_i = sheet[f'{d["остаток Иркутск"]}{i}'].value
                            except:
                                edge_i = None

                            if edge_m:
                                comm = f'{comm}, Москва'
                            if edge_n:
                                comm = f'{comm}, Новосибирск'
                            if edge_i:
                                comm = f'{comm}, Иркутск'
                            if comm:
                                edge = None
                                status = f'В наличии в: {comm}'
                except:
                    try:
                        edge = sheet[f'{d["свободно"]}{i}'].value
                        print(f'Свободные {edge}')
                        if edge and int(price_validator(edge)) > 0:
                            status = 'В наличии'
                    except:
                        status = ''
                        comm = None
                        try:
                            edge_m = sheet[f'{d["остаток Москва"]}{i}'].value
                        except:
                            edge_m = None
                        try:
                            edge_n = sheet[f'{d["остаток Новос"]}{i}'].value
                        except:
                            edge_n = None
                        try:
                            edge_i = sheet[f'{d["остаток Иркутск"]}{i}'].value
                        except:
                            edge_i = None

                        if edge_m:
                            comm = f'{comm}, Москва'
                        if edge_n:
                            comm = f'{comm}, Новосибирск'
                        if edge_i:
                            comm = f'{comm}, Иркутск'
                        if comm:
                            edge = None
                            status = f'В наличии в: {comm}'

                print(status)
                print(edge)
                ok = loop.run_until_complete(unique_prods(name))
                if (max_ == 'Н.Д' and min_ == 'Н.Д') or not (max_ and min_):
                    continue

                pic = 'Н.Д.'

                print(f'Товар {name} прошел проверку на просто заголовок', min_, max_)

                if not status:
                    if edge and edge != '' and edge != ' ':
                        d_edge = ''.join(filter(str.isdigit, str(edge)))
                        try:
                            if int(d_edge) > 0:
                                status = 'В наличии'
                        except:
                            status = 'Нет в наличии'
                            edge = 0
                    else:
                        status = 'Нет в наличии'
                        edge = 0

                if not status:
                    status = 'Н.Д.'
                else:
                    try:
                        m = sheet[f'{d["остаток Москва"]}{i}'].value
                    except:
                        m = 0
                    try:
                        n = sheet[f'{d["остаток Новос"]}{i}'].value
                    except:
                        n = 0
                    try:
                        ii = sheet[f'{d["остаток Иркутск"]}{i}'].value
                    except:
                        ii = 0

                    if edge_m or edge_i or edge_n:
                        descr = f'{descr} {status}, Москва: {m}шт, Иркутск: {n}шт, Новосибирск: {ii}шт'
                    else:
                        descr = f'{descr} {status}. Осталось товара - {edge}'

                    if descr:
                        chars = chat_get_chars(name, f'{descr}')
                    else:
                        chars = chat_get_chars(name, name)

                    if descr:
                        type_ = chat_get_type(name, descr)
                    else:
                        type_ = chat_get_type(name, name)
                chars_p = {
                    'name': name, 'art': art, 'descr': descr, 'pic': pic, 'min_': min_, 'max_': max_, 'edge': edge,
                    'chars': chars, 'type_': type_, 'available': status
                }
            count += 1
            if ok != False:
                update_exist_prod_file(chars_p, name)
                continue

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(add_prod_from_sites(chars, chars_p=chars_p, url=provider.name))

            his = History.objects.create(i_c=user,
                                         done=f'Добавился прайс этого поставщика - {provider.name}, добавлено или обновлено было {count} товаров.',
                                         date_of_change=datetime.today(), title_prod='Загрузка прайса')

    elif 'csv' in name_f:
        values = list(d.values())
        if not values[0].isdigit():
            d = {k: column_index_from_string(v)-1 for k, v in d.items() if v != 'Н.Д.'}

        with open(file_path, newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                print(f'Парсим новый товар номером {reader.line_num}')
                try:
                    name = row[reader.fieldnames[d["имя товара"]]]
                except:
                    continue

                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                if name is None or name == 'имя товара':
                    continue

                try:
                    art = row[reader.fieldnames[d["артикул"]]]
                except:
                    art = 'Н.Д.'

                try:
                    descr = row[reader.fieldnames[d["описание"]]]
                except:
                    descr = 'Н.Д.'

                try:
                    min_ = price_validator(row[reader.fieldnames[d["опт мин"]]])
                    try:
                        min_ = min_ - ((min_ / 100) * price_validator(row[reader.fieldnames[d["скидка"]]]))
                        if not min_:
                            min_ = price_validator(row[reader.fieldnames[d["опт мин"]]])
                    except:
                        pass
                except:
                    min_ = 'Н.Д.'

                try:
                    max_ = price_validator(row[reader.fieldnames[d["опт макс"]]])
                    try:
                        max_ = max_ - ((max_ / 100) * price_validator(row[reader.fieldnames[d["скидка"]]]))
                        if not max_:
                            max_ = price_validator(row[reader.fieldnames[d["опт макс"]]])
                    except:
                        pass
                except:
                    max_ = 'Н.Д.'

                status = None
                edge_m = None
                edge_n = None
                edge_i = None
                edge = None
                if d.get("остаток Москва", None):
                    edge_m = row[reader.fieldnames[d["остаток Москва"]]]
                    print(f'Москва {edge_m}')

                try:
                    edge = row[d["остаток на складе"]]
                    if edge and int(price_validator(edge)) > 0:
                        status = 'В наличии'
                    else:
                        edge = row[d["свободно"]]
                        if edge and int(price_validator(edge)) > 0:
                            status = 'В наличии'
                        else:
                            status = ''
                            comm = None
                            try:
                                edge_m = row[reader.fieldnames[d["остаток Москва"]]]
                            except:
                                edge_m = None
                            try:
                                edge_n = row[reader.fieldnames[d["остаток Новос"]]]
                            except:
                                edge_n = None
                            try:
                                edge_i = row[reader.fieldnames[d["остаток Иркутск"]]]
                            except:
                                edge_i = None

                            if edge_m:
                                comm = f'{comm}, Москва'
                            if edge_n:
                                comm = f'{comm}, Новосибирск'
                            if edge_i:
                                comm = f'{comm}, Иркутск'
                            if comm:
                                edge = None
                                status = f'В наличии в: {comm}'
                except:
                    try:
                        edge = row[d["свободно"]]
                        if edge and int(price_validator(edge)) > 0:
                            status = 'В наличии'
                    except:
                        status = ''
                        comm = None
                        try:
                            edge_m = row[reader.fieldnames[d["остаток Москва"]]]
                        except:
                            edge_m = None
                        try:
                            edge_n = row[reader.fieldnames[d["остаток Новос"]]]
                        except:
                            edge_n = None
                        try:
                            edge_i = row[reader.fieldnames[d["остаток Иркутск"]]]
                        except:
                            edge_i = None

                        edge = 0

                        if edge_m:
                            try:
                                edge += price_validator(edge_m)
                                comm = f'{comm}, Москва'
                            except:
                                pass
                        if edge_n:
                            try:
                                edge += price_validator(edge_n)
                                comm = f'{comm}, Новосибирск'
                            except:
                                pass
                        if edge_i:
                            try:
                                edge += price_validator(edge_i)
                                comm = f'{comm}, Иркутск'
                            except:
                                pass
                        if comm:
                            edge = int(edge)
                            status = f'В наличии в: {comm}'

                ok = loop.run_until_complete(unique_prods(name))
                if (max_ == 'Н.Д' and min_ == 'Н.Д') or not (max_ and min_):
                    continue

                pic = 'Н.Д.'

                print(f'Товар {name} прошел проверку на просто заголовок', min_, max_)

                if descr:
                    chars = chat_get_chars(name, f'{descr}')
                else:
                    chars = chat_get_chars(name, name)

                print(chars)

                if descr:
                    type_ = chat_get_type(name, descr)
                else:
                    type_ = chat_get_type(name, name)

                print(type_)
                if not status:
                    if edge and edge != '' and edge != ' ':
                        d_edge = ''.join(filter(str.isdigit, str(edge)))
                        try:
                            if int(d_edge) > 0:
                                status = 'В наличии'
                        except:
                            status = 'Нет в наличии'
                            edge = 0
                    else:
                        status = 'Нет в наличии'
                        edge = 0

                if not status:
                    status = 'Н.Д.'
                else:
                    try:
                        m = row[reader.fieldnames[d["остаток Москва"]]]
                    except:
                        m = 0
                    try:
                        n = row[reader.fieldnames[d["остаток Новос"]]]
                    except:
                        n = 0
                    try:
                        ii = row[reader.fieldnames[d["остаток Иркутск"]]]
                    except:
                        ii = 0

                    if edge_m or edge_i or edge_n:
                        descr = f'{descr} {status}, Москва: {m}шт, Иркутск: {n}шт, Новосибирск: {ii}шт'
                    else:
                        descr = f'{descr} {status}. Осталось товара - {edge}'

                chars_p = {
                    'name': name, 'art': art, 'descr': descr, 'pic': pic, 'min_': min_, 'max_': max_, 'edge': edge,
                    'chars': chars, 'type_': type_, 'available': status
                }

                count += 1
                if ok != False:
                    update_exist_prod_file(chars_p, url=provider.name)
                    continue

                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                print("Пошло добавление файла")
                loop.run_until_complete(add_prod_from_sites(chars, chars_p=chars_p, url=provider.name))


pars_file(parameters[0], parameters[1], parameters[2])

